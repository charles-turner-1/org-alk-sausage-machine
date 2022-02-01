#########################
#IMPORT REQUIRED PACKAGES
#########################

import numpy as np
import pandas as pd
import math
def truncate(number, digits) -> float:
    stepper = 10.0 ** digits
    return math.trunc(stepper * number) / stepper
from scipy.stats import linregress
from lmfit import Minimizer, Parameters, report_fit
import matplotlib.pyplot as plt
from matplotlib.legend_handler import HandlerBase
import os
from openpyxl import load_workbook
from IPython.display import Markdown, display
def printmd(string):
    display(Markdown(string))
class MarkerHandler(HandlerBase):
    def create_artists(self, legend, tup,xdescent, ydescent,
                        width, height, fontsize,trans):
        return [plt.Line2D([width/2], [height/2.],ls="",
                       marker=tup[1],color=tup[0], transform=trans)]
    



#######################################################
#IMPORT PROCESSED EXCEL FILES CONTAINING TITRATION DATA
#######################################################

#FIRST TITRATION (TA TITRATION)
df_TA = pd.read_excel("/Users/dankerr/Desktop/DESKTOP/POSTGRAD/MASTER SEAWATER ANALYSIS/CARBON/Alkalinity Titration/BACK TITRATION/SODIUM ACETATE/PROCESSED_NACO3COO/01.09.21.50UM.001_PROCESSED.xlsx")
#RETURN SAMPLE TO ORIGINAL pH USING NaOH TITRATION
df_NaOH = pd.read_excel("/Users/dankerr/Desktop/DESKTOP/POSTGRAD/MASTER SEAWATER ANALYSIS/CARBON/Alkalinity Titration/BACK TITRATION/SODIUM ACETATE/PROCESSED_NACO3COO/01.09.21.50UM.001.NAOH_PROCESSED.xlsx")
#SECOND TITRATION (BACK TITRATION)
df_BT = pd.read_excel("/Users/dankerr/Desktop/DESKTOP/POSTGRAD/MASTER SEAWATER ANALYSIS/CARBON/Alkalinity Titration/BACK TITRATION/SODIUM ACETATE/PROCESSED_NACO3COO/01.09.21.50UM.001.BT_PROCESSED.xlsx")

####################################################################
##DEFINE FUNCTION TO SAVE PROCESSED TA DATA TO TA MASTER SPREADSHEET
####################################################################


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2', 
                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name, 
            startrow=startrow if startrow is not None else 0, 
            **to_excel_kwargs)
        return
    
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)
    
    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)
    
    # copy existing sheets
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
    



##############################################################################################################################
##############################################################################################################################
################################## FIRST TITRATION (TA TITRATION) ############################################################
##############################################################################################################################
##############################################################################################################################


##################
#ACID TITRANT INFO
##################
C = 0.10060392 # Estimate of OVERALL HCl titrant concentration (mol.kg-1) from CRM_182_0392 
Cl_HCl = 0.10060392 #ionic strength of acid, 0.1M HCl made up in DI therefore Cl-] = 0.1 M

#######################################################
#CONSTANTS CALCULATED BASED ON SALINITY AND TEMPERATURE 
#######################################################
R = 8.314472 # Universal gas constant
F = 96485.3399 # Faraday constant

################################
#EXTRACT DATA FROM TA EXCEL FILE
################################
V0 = df_TA.iloc[0]['g_0']-df_TA.iloc[0]['g_1'] # Sample mass (g)
S_TA = df_TA.iloc[0]['SALINITY']  # Sample Salinity 
sample_id_TA = df_TA.iloc[0]['SAMPLE']  # Sample ID
data_start_TA = int(df_TA.iloc[0]['data_start']-1) #row which titration starts, eg after initial acid addition and degassing
initial_EV_TA = df_TA.iloc[9]['102 Voltage (V)'] #EV of sample before any acid addition, at index = 10



#######################################################################################################################
#strip data from df such that index = 1 aligns with start of titration eg after initial acid addition and degassing
#######################################################################################################################
df_TA['E(V)'] = df_TA.drop(df_TA.index[0:data_start_TA],axis=0)['102 Voltage (V)']
df_TA['E(V)'] = df_TA['E(V)'].shift(-(data_start_TA))
df_TA['Sample_T'] = df_TA.drop(df_TA.index[0:data_start_TA],axis=0)['SAMPLE Temperature (°C)']
df_TA['Sample_T'] = df_TA['Sample_T'].shift(-(data_start_TA))
df_TA['Acid_T'] = df_TA.drop(df_TA.index[0:data_start_TA],axis=0)['ACID Temperature (°C)']
df_TA['Acid_T'] = df_TA['Acid_T'].shift(-(data_start_TA))
stripped_df_TA = df_TA[['E(V)', 'Sample_T', 'Acid_T', "mL"]].copy() #copy above variables to new "stripped_df"
stripped_df_TA = stripped_df_TA.dropna() #remove NaN values from stripped_df
df_TA = stripped_df_TA #for congruency with existing code, redefine stripped_df as df
 

#####################################
#SAMPLE TEMPERATURE AND NERNST FACTOR
#####################################
T_TA = df_TA["T"] = (df_TA["Sample_T"]+273.15)  # CREATE COLUMN SAMPLE TEMPERATURE (KELVIN) AT EACH TITRATION POINT
df_TA["K"] = (R*T_TA)/F # Nernst factor 
initial_K_TA = df_TA.iloc[9]['K'] # Initial Nernst factor, used to calculate initial pH


###################################################
#ACID DENSITY, CONVERSION FROM VOLUME TO MASS UNITS
###################################################
df_TA["rho"] = (df_TA["Acid_T"]*-0.0008958+1.02119193) #  CREATE COLUMN OF DENSITY (g.cm-3) OF ACID AT EACH TITRATION POINT
initial_mL_acid = df_TA.iloc[0]["mL"] #define value for initial volume acid added (mL)
intial_g_acid = initial_mL_acid*df_TA.iloc[0]["rho"]#convert inital volume of acid added to mass value (g)
df_TA['delta_mL'] = df_TA['mL'].diff() #calculate the incremental values of volume of acid added
df_TA['delta_g'] = df_TA['delta_mL']*df_TA["rho"] #convert the incremental values of volume of acid added to mass values
df_TA = df_TA.fillna(0) #initial value of df['delta_g'] will be NA by default, replace with 0
df_TA['m'] = intial_g_acid+np.cumsum(df_TA['delta_g'])#caculate cumulative total of mass of acid added, (intial mass of acid added + incremment mass 1,  intial mass of acid added + incremment mass 1 + incremment mass 2 ...)

############################
#IONIC STRENGTH AND SALINITY 
############################
ImO_TA = (19.924*S_TA/(1000-1.005*S_TA)) # ionic strength from Dickson 2007
df_TA['ImO'] = (ImO_TA*V0+df_TA.m*(Cl_HCl))/(V0+df_TA.m) # ionic strength adjusted for addition of acid
df_TA["S"] = (1000*df_TA.ImO)/(19.924+1.005*df_TA.ImO) # salinity adjusted for addition of acid


############################################################
#EQUILIBRIUM CONSTANTS, BISULFATE ION AND HYDROGEN FLUORIDE  
############################################################
df_TA['K_S'] = np.exp(-4276.1/T_TA + 141.328 - 23.093*np.log(T_TA)+(-13856/T_TA + 324.57- 47.986*np.log(T_TA))*(ImO_TA**(1/2)) +(35474/T_TA  - 771.54 +114.723*np.log(T_TA))*ImO_TA - (2698/T_TA)*(ImO_TA)**(3/2)+(1776/T_TA)*(ImO_TA)**(2) +np.log(1-0.001005*S_TA)) # pKa Bisulfate ion [HSO4-] K_S from Dickson1990
df_TA['K_F'] = np.exp(874/T_TA - 9.68 + 0.111*S_TA**0.5)  # pKa Hydrogen Fluoride ion [HF] K_F  from Perex & Fraga 1987

#####################################
#TOTAL SULPHATE AND HYDROGEN FLUORIDE 
#####################################
df_TA['S_T'] =  (0.14/96.062)*(df_TA.S/1.80655)# Total Sulphate concentration S_T from Morris & Riley 1966
df_TA['F_T'] = (0.000067/18.998)*(df_TA.S/1.80655) # Total Hydrogen Fluoride concentration F_T from Riley 1996
df_TA["Z"] = (1+df_TA.S_T/df_TA.K_S) #from dickson 2007 Z = (1+S_T/K_S) 

################################################################
#GRAN FUNCTION, IDENTIFY TA INITIAL ESTIMATE AND E0, INITIAL GRAN [H'] AND GRAN pH
################################################################
df_TA["F1"] = ((V0+df_TA["m"])*np.exp((df_TA["E(V)"]/(df_TA.K)))) #Calculate Gran Function F1 at each titration point
df_TA = df_TA[df_TA["F1"].between(10000, 1000000)] #drop all gran funciton values less than 10000 as these are TYPICALLY non linear with respect to m
slope, intercept, r_value, p_value, std_err =linregress(df_TA["m"], df_TA["F1"])#CALL SLOPE AND INTERCEPT OF Gran function F1
equivalence_point_TA = -intercept/slope #Calculate equivalnce point estimate (g) from Gran function F1
TA_est_TA = (equivalence_point_TA*C)/V0#Estimate TA using equivalnce point estimate

df_TA["E0_est"] = (df_TA["E(V)"]-(df_TA["K"])*np.log((-V0*TA_est_TA + df_TA["m"]*C)/(V0 + df_TA["m"]))) #CALCULATE EO ESTIMATE FOR EACH TITRATION POINT
E0_init_est_TA = df_TA["E0_est"].mean()#AVERAGE ALL EO VALUES TO OBTAIN AN INITIAL ESTIMATE OF EO
df_TA["H"] = (np.exp((df_TA["E(V)"]-E0_init_est_TA)/(df_TA["K"]))) #USING EO INITIAL ESTIMATE CALCULATE [H'] FOR EACH TITRATION POINT
df_TA["GRAN_pH"] = -(np.log10(np.exp((df_TA["E(V)"]-E0_init_est_TA)/df_TA["K"])))#CALCULATE GRAN pH


##########################################
#TA EQUATION TO DEFINE VECTOR OF RESIDUALS
##########################################
f = 1 #define f and provide inital estimate
df_TA["residuals"] = TA_est_TA  + ((V0+df_TA.m)/V0)*((f*df_TA.H)/df_TA.Z) -(df_TA.m/V0)*C # DEFINE RESIDUALS
SSR = (np.sum((TA_est_TA + ((V0+df_TA.m)/V0)*((f*df_TA.H)/df_TA.Z) -(df_TA.m/V0)*C )**2))*10**12 # DEFINE SSR EQUATION


#################################################
#NON LINEAR LEAST SQUARES CURVE FITTING PROCEDURE
#################################################

new_df_TA = df_TA[df_TA["GRAN_pH"].between(3, 3.5)]#SELECT ONLY TITRATION POINTS WHICH pH ARE BETWEEN 3.0 - 3.5 
data_points_TA = len(new_df_TA.index) #CALCULATE NUMBER OF DATA POINTS IN 3-3.5 pH RANGE

# DEFINE FUNCTION WHICH RETURNS VALUE(S) TO BE MINIMISED, IN THIS CASE SSR 
x_TA = new_df_TA.H
data_TA = new_df_TA.m

def fcn2min(params, x_TA, data_TA):
    f_NLSF_TA = params['f_NLSF_TA']
    TA_est_NLSF_TA = params['TA_est_NLSF_TA']
    model = ((np.sum((TA_est_NLSF_TA + 
                      ((V0+new_df_TA.m)/V0)*
                      ((f_NLSF_TA*new_df_TA.H)/new_df_TA.Z) -
                      (new_df_TA.m/V0)*C)**2))*10**12)
    
    return model - data_TA

# create a set of Parameters
params = Parameters()
params.add('f_NLSF_TA',   value= 1)
params.add('TA_est_NLSF_TA', value= TA_est_TA)
# do fit, here with leastsq model
minner = Minimizer(fcn2min, params, fcn_args=(x_TA, data_TA))
kws  = {'options': {'maxiter':10}}
result = minner.minimize()
# calculate final result
final = data_TA + result.residual


#################################
#EXTRACT AND PROCESS TA NLSF RESULTS
#################################
TA_processed_TA = result.params.get('TA_est_NLSF_TA').value #EXTRACT INTIAL TA VALUE M/KG-1
TA_final_TA = TA_processed_TA*10**6 #COVERT INTIAL TA VALUE TO µmol/KG-1
f_TA = result.params.get('f_NLSF_TA').value #EXTRACT NLSF F VALUE
E0_processed_TA = E0_init_est_TA + df_TA["K"]*np.log(f) #CALCULATE E0 FROM NLSF F VALUE
E0_final_TA = E0_processed_TA.mean() #FINAL ESTIMATE OF E0
new_df_TA["pH"] = -np.log10(np.exp((df_TA["E(V)"]-E0_final_TA)/df_TA["K"])) #CALCULATE pH AT EACH TITRATION POINT FROM E0 FINAL ESTIMATE
initial_pH_TA = -np.log10(np.exp((initial_EV_TA-E0_final_TA)/initial_K_TA)) #CALCULATE INTIAL pH FROM E0 FINAL ESTIMATE







##############################################################################################################################
##############################################################################################################################
################################## NAOH TITRATION ############################################################################
##############################################################################################################################
##############################################################################################################################

##################
#BASE TITRANT INFO
##################
C_NaOH = 0.082744091 # ± 0.000226775 determimed from acidic gran function of standardisation of NaOH using crm standardised HCl
I_NaOH = 0.082744091 #ionic strength of NaOH 

################################
#EXTRACT DATA FROM TA EXCEL FILE
################################
df_NaOH["T"] = (df_NaOH["SAMPLE Temperature (°C)"]+273.15) #create colume for sample temperature (KELVIN) at each titration point
df_NaOH["NaOH_T"] = df_NaOH["NaOH Temperature (°C)"] #create colume for temperature (Degrees Celsius) of NaOH upon addition to cell 

###################################################
#BASE DENSITY, CONVERSION FROM VOLUME TO MASS UNITS
###################################################
df_NaOH["rho"] = (df_NaOH["NaOH_T"]*-0.014702658+1.27068345) # Density of NaOH titrant g.cm-3 
initial_mL_base = 0 #define value for initial volume base added (mL)
intial_g_base = initial_mL_base*df_NaOH.iloc[0]["rho"]#convert inital volume of base added to mass value (g)
df_NaOH['delta_mL'] = df_NaOH['mL'].diff() #calculate the incremental values of volume of base added
df_NaOH['delta_g'] = df_NaOH['delta_mL']*df_NaOH["rho"] #convert the incremental values of volume of base added to mass values
#df_NaOH = df_NaOH.fillna(0) #initial value of df['delta_g'] will be NA by default, replace with 0
df_NaOH['delta_g'] = df_NaOH['delta_g'].fillna(0)
df_NaOH['m'] = intial_g_base+np.cumsum(df_NaOH['delta_g'])#caculate cumulative total of mass of base added, (intial mass of acid added + incremment mass 1,  intial mass of acid added + incremment mass 1 + incremment mass 2 ...)

############################
#IONIC STRENGTH AND SALINITY 
############################
Va = df_TA['m'][df_TA.index[-1]] #DEFINE TOTAL MASS OF ACID ADDED DURING FIRST (TA) TITRATION
Vb = df_NaOH['m'][df_NaOH.index[-1]] #DEFINE TOTAL MASS OF BASE ADDED DURING NAOH TITRATION

S_NaOH_initial = df_TA['S'][df_TA.index[-1]] #INITIAL SALINITY VALUE FOR NAOH TITRATION DEFINED AS LAST salinty value of FIRST titration (TA TITRATION), which factors in slight dilution due to addition of acid
ImO_NaOH_initial = (19.924*S_NaOH_initial/(1000-1.005*S_NaOH_initial)) #calculate INTIAL ImO from this S value
df_NaOH['ImO_NaOH'] = (ImO_NaOH_initial*(V0+Va)+(df_NaOH['m']*I_NaOH))/(V0 + Va + df_NaOH['m']) # FINAL IMO VALUE AT END OF NAOH TITRATION
df_NaOH['S'] = (1000*df_NaOH['ImO_NaOH'])/(19.924+1.005*df_NaOH['ImO_NaOH']) # FINAL SALINITY VALUE AT END OF NAOH TITRATION
S_NaOH_final = df_NaOH['S'][df_NaOH.index[-1]]


###############################
#CALCULATE pH, [H+], [OH-], H0
###############################
df_NaOH["K"] = (R*df_NaOH["T"])/F #Get K value at eachpoint during NaOH titration
df_NaOH["pH"] = -np.log10(np.exp((df_NaOH["102 Voltage (V)"]-E0_final_TA)/df_NaOH["K"]))#Using EO estimated from TA NLSF procedure calculate pH at each point during NaOH titration
df_NaOH["H"] = 10**-(df_NaOH["pH"]) #Using pH calculate [H+] at each titration point
df_NaOH['pKw'] = (-np.log10(np.exp(148.9652-13847.26/df_NaOH["T"]-23.6521*np.log(df_NaOH["T"])+(-5.977+118.67/df_NaOH["T"]+1.0495*np.log(df_NaOH["T"]))*df_NaOH["S"]**0.5-0.01615*df_NaOH["S"]))) #Acid dissociation constant of Water
df_NaOH['OH'] = (10**-df_NaOH["pKw"])/df_NaOH["H"] #using Acid dissociation constant of Water and [H+] calculate [OH-]
initial_EV_NaOH = df_NaOH.iloc[0]['102 Voltage (V)'] #Proton concentration prior to NaOH addition, H0, defined as [H+] at end of first (TA) titration (start of NaOH titration)
initial_K_NaOH = df_NaOH.iloc[0]['K']
H0 = (np.exp((initial_EV_NaOH-E0_final_TA)/(initial_K_NaOH)))






##############################################################################################################################
##############################################################################################################################
################################## SECOND TITRATION (BT TITRATION) ###########################################################
##############################################################################################################################
##############################################################################################################################


################################
#EXTRACT DATA FROM TA EXCEL FILE
################################
V0_BT = (V0+Va+Vb) # Sample mass accounting for additions of acid and base (g) 
sample_id_BT = df_BT.iloc[0]['SAMPLE']  # Sample ID
data_start_BT = int(df_BT.iloc[0]['data_start']-1) #row which titration starts, eg after initial acid addition and degassing


#######################################################################################################################
#strip data from df such that index = 1 aligns with start of titration eg after initial acid addition and degassing
#######################################################################################################################
df_BT['E(V)'] = df_BT.drop(df_BT.index[0:data_start_BT],axis=0)['102 Voltage (V)']
df_BT['E(V)'] = df_BT['E(V)'].shift(-(data_start_BT))
df_BT['Sample_T'] = df_BT.drop(df_BT.index[0:data_start_BT],axis=0)['SAMPLE Temperature (°C)']
df_BT['Sample_T'] = df_BT['Sample_T'].shift(-(data_start_BT))
df_BT['Acid_T'] = df_BT.drop(df_BT.index[0:data_start_BT],axis=0)['ACID Temperature (°C)']
df_BT['Acid_T'] = df_BT['Acid_T'].shift(-(data_start_BT))
#copy above variables to new "stripped_df_BT"
stripped_df_BT = df_BT[['E(V)', 'Sample_T', 'Acid_T', "mL"]].copy()
#remove NaN values from stripped_df_BT
stripped_df_BT = stripped_df_BT.dropna()
#for congruency with existing code, redefine stripped_df_BT as df_BT
df_BT = stripped_df_BT

###################################################
#ACID DENSITY, CONVERSION FROM VOLUME TO MASS UNITS
###################################################
df_BT["rho"] = (df_BT["Acid_T"]*-0.0008958+1.02119193) #  CREATE COLUMN OF DENSITY (g.cm-3) OF ACID AT EACH TITRATION POINT
initial_mL_acid = df_BT.iloc[0]["mL"] #define value for initial volume acid added (mL)
intial_g_acid = initial_mL_acid*df_BT.iloc[0]["rho"]#convert inital volume of acid added to mass value (g)
df_BT['delta_mL'] = df_BT['mL'].diff() #calculate the incremental values of volume of acid added
df_BT['delta_g'] = df_BT['delta_mL']*df_BT["rho"] #convert the incremental values of volume of acid added to mass values
df_BT = df_BT.fillna(0) #initial value of df['delta_g'] will be NA by default, replace with 0
df_BT['m'] = intial_g_acid+np.cumsum(df_BT['delta_g'])#caculate cumulative total of mass of acid added, (intial mass of acid added + incremment mass 1,  intial mass of acid added + incremment mass 1 + incremment mass 2 ...)

#####################################
#SAMPLE TEMPERATURE AND NERNST FACTOR
#####################################
T_BT = df_BT["T"] = (df_BT["Sample_T"]+273.15)  # CREATE COLUMN SAMPLE TEMPERATURE (KELVIN) AT EACH TITRATION POINT
df_BT["K"] = (R*T_BT)/F # Nernst factor 
initial_K_BT = df_BT.iloc[0]['K'] # Initial Nernst factor, used to calculate initial pH
initial_EV_BT = df_BT.iloc[0]['E(V)'] #EV of sample before any acid addition

############################
#IONIC STRENGTH AND SALINITY 
############################
S_BT = S_NaOH_final #intial Salinity value of second (BT) titration is defined as last Salinity value of NaOH titration
ImO_BT = (19.924*S_BT/(1000-1.005*S_BT)) # ionic strength from Dickson 2007
df_BT['ImO'] = (ImO_BT*V0_BT+df_BT.m*(Cl_HCl))/(V0_BT+df_BT.m) # ionic strength adjusted for addition of acid
df_BT["S"] = (1000*df_BT.ImO)/(19.924+1.005*df_BT.ImO) # salinity adjusted for addition of acid


############################################################
#EQUILIBRIUM CONSTANTS, BISULFATE ION AND HYDROGEN FLUORIDE  
############################################################
df_BT['K_S'] = np.exp(-4276.1/T_BT + 141.328 - 23.093*np.log(T_BT)+(-13856/T_BT + 324.57- 47.986*np.log(T_BT))*(ImO_BT**(1/2)) +(35474/T_BT  - 771.54 +114.723*np.log(T_BT))*ImO_BT - (2698/T_BT)*(ImO_BT)**(3/2)+(1776/T_BT)*(ImO_BT)**(2) +np.log(1-0.001005*S_BT)) # pKa Bisulfate ion [HSO4-] K_S from Dickson1990
df_BT['K_F'] = np.exp(874/T_BT - 9.68 + 0.111*S_BT**0.5)  # pKa Hydrogen Fluoride ion [HF] K_F  from Perex & Fraga 1987

#####################################
#TOTAL SULPHATE AND HYDROGEN FLUORIDE 
#####################################
df_BT['S_T'] =  (0.14/96.062)*(df_BT.S/1.80655)# Total Sulphate concentration S_T from Morris & Riley 1966
df_BT['F_T'] = (0.000067/18.998)*(df_BT.S/1.80655) # Total Hydrogen Fluoride concentration F_T from Riley 1996
df_BT["Z"] = (1+df_BT.S_T/df_BT.K_S) #from dickson 2007 Z = (1+S_T/K_S) 

################################################################
#GRAN FUNCTION, IDENTIFY TA INITIAL ESTIMATE AND E0, INITIAL GRAN [H'] AND GRAN pH
################################################################
df_BT["F1"] = ((V0_BT+df_BT["m"])*np.exp((df_BT["E(V)"]/(df_BT.K)))) #Calculate Gran Function F1 at each titration point
df_BT = df_BT[df_BT["F1"].between(10000, 1000000)] #drop all gran funciton values less than 10000 as these are TYPICALLY non linear with respect to m
slope, intercept, r_value, p_value, std_err =linregress(df_BT["m"], df_BT["F1"])#CALL SLOPE AND INTERCEPT OF Gran function F1
equivalence_point_BT = -intercept/slope #Calculate equivalnce point estimate (g) from Gran function F1
TA_est_BT = (equivalence_point_BT*C)/V0_BT#Estimate TA using equivalnce point estimate

df_BT["E0_est"] = (df_BT["E(V)"]-(df_BT["K"])*np.log((-V0_BT*TA_est_BT + df_BT["m"]*C)/(V0_BT + df_BT["m"]))) #CALCULATE EO ESTIMATE FOR EACH TITRATION POINT
E0_init_est_BT = df_BT["E0_est"].mean()#AVERAGE ALL EO VALUES TO OBTAIN AN INITIAL ESTIMATE OF EO
df_BT["H"] = (np.exp((df_BT["E(V)"]-E0_init_est_BT)/(df_BT["K"]))) #USING EO INITIAL ESTIMATE CALCULATE [H'] FOR EACH TITRATION POINT
df_BT["GRAN_pH"] = -(np.log10(np.exp((df_BT["E(V)"]-E0_init_est_BT)/df_BT["K"])))#CALCULATE GRAN pH


##########################################
#TA EQUATION TO DEFINE VECTOR OF RESIDUALS
##########################################
f = 1 #define f and provide inital estimate
df_BT["residuals"] = TA_est_BT  + ((V0_BT+df_BT.m)/V0_BT)*((f*df_BT.H)/df_BT.Z) -(df_BT.m/V0_BT)*C # DEFINE RESIDUALS
SSR = (np.sum((TA_est_BT + ((V0_BT+df_BT.m)/V0_BT)*((f*df_BT.H)/df_BT.Z) -(df_BT.m/V0_BT)*C )**2))*10**12 # DEFINE SSR EQUATION


#################################################
#NON LINEAR LEAST SQUARES CURVE FITTING PROCEDURE
#################################################

new_df_BT = df_BT[df_BT["GRAN_pH"].between(3, 3.5)]#SELECT ONLY TITRATION POINTS WHICH pH ARE BETWEEN 3.0 - 3.5 
data_points_BT = len(new_df_BT.index) #CALCULATE NUMBER OF DATA POINTS IN 3-3.5 pH RANGE

# DEFINE FUNCTION WHICH RETURNS VALUE(S) TO BE MINIMISED, IN THIS CASE SSR 
x_BT = new_df_BT.H
data_BT = new_df_BT.m

def fcn2min(params, x_BT, data_BT):
    f_NLSF_BT = params['f_NLSF_BT']
    TA_est_NLSF_BT = params['TA_est_NLSF_BT']
    model = ((np.sum((TA_est_NLSF_BT + 
                      ((V0_BT+new_df_BT.m)/V0_BT)*
                      ((f_NLSF_BT*new_df_BT.H)/new_df_BT.Z) -
                      (new_df_BT.m/V0_BT)*C)**2))*10**12)
    
    return model - data_BT

# create a set of Parameters
params = Parameters()
params.add('f_NLSF_BT',   value= 1)
params.add('TA_est_NLSF_BT', value= TA_est_BT)
# do fit, here with leastsq model
minner = Minimizer(fcn2min, params, fcn_args=(x_BT, data_BT))
kws  = {'options': {'maxiter':10}}
result = minner.minimize()
# calculate final result
final = data_BT + result.residual


#################################
#EXTRACT AND PROCESS TA NLSF RESULTS
#################################
TA_processed_BT = result.params.get('TA_est_NLSF_BT').value #EXTRACT INTIAL TA VALUE M/KG-1
TA_final_BT = TA_processed_BT*10**6 #COVERT INTIAL TA VALUE TO µmol/KG-1
f_BT = result.params.get('f_NLSF_BT').value #EXTRACT NLSF F VALUE
E0_processed_BT = E0_init_est_BT + df_BT["K"]*np.log(f_BT) #CALCULATE E0 FROM NLSF F VALUE
E0_final_BT = E0_processed_BT.mean() #FINAL ESTIMATE OF E0
new_df_BT["pH"] = -np.log10(np.exp((df_BT["E(V)"]-E0_final_BT)/df_BT["K"])) #CALCULATE pH AT EACH TITRATION POINT FROM E0 FINAL ESTIMATE
initial_pH_BT = -np.log10(np.exp((initial_EV_BT-E0_final_BT)/initial_K_BT)) #CALCULATE INTIAL pH FROM E0 FINAL ESTIMATE

print('TA: ', TA_final_TA, 'µmol.kg-1')
print('OrgAlk: ', TA_final_BT, 'µmol.kg-1')[
