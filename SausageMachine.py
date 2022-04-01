#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun Jan 23 16:37:58 2022

@author: Dan Kerr & Charles Turner

This file will comprise the main class used in the sausage machine to perform
organic alkalinity calculations.
"""

from asyncio import sslproto
from multiprocessing.sharedctypes import Value
from re import M
import numpy as np
import pandas as pd
import math
import os
import matplotlib.pyplot as plt

from scipy.stats import linregress
from lmfit import Minimizer, Parameters, report_fit
from matplotlib.legend_handler import HandlerBase
from openpyxl import load_workbook
from IPython.display import Markdown, display


class org_alk_titration():
    def __init__(self,dataset_path=None,spreadsheet_name_TA = None
                                  ,spreadsheet_name_NaOH = None
                                  ,spreadsheet_name_OA = None):
        self.dataset_path = dataset_path
        self.spreadsheet_name_TA = spreadsheet_name_TA
        self.spreadsheet_name_NaOH = spreadsheet_name_NaOH
        self.spreadsheet_name_OA = spreadsheet_name_OA
        self.S_TA = None
        self.V0 = None
        self.df_TA = None
        self.df_NaOH = None
        self.df_OA = None
        self.titration_features["TA"]["mass_known"] = False
        self.titration_features["NaOH"]["mass_known"] = False
        self.temp_TA_known = False
        self.E0_init_est_TA = None
        self.E0_init_est_OA = None
    ########################################################
    #CONSTANTS CALCULATED BASED ON SALINITY AND TEMPERATURE#
    ########################################################

    R = 8.314472 # Universal gas constant
    F = 96485.3399 # Faraday constant


    titration_features = {
        "TA" : {
            "slope_rho" : -0.0008958,
            "intercept_rho" : 1.02119193,
            "df_T_label" : "Acid_T",   
            "mass_known" : False,
            "TA_est" : None,
            "E0_init_est" : None,
            "initial_EV" : None,
            "initial_K" : None,
            "titration_soln" : None,
            "E0_final" : None,
            "TA_final" : None,
            "TA_processed" : None,
        },
        "NaOH" : {
            "slope_rho" : -0.014702658,
            "intercept_rho" : 1.27068345,
            "df_T_label" : "NaOH_T",
            "mass_known" : False,
            "TA_est" : None,
            "E0_init_est" : None,
            "initial_EV" : None,
            "initial_K" : None,
            "titration_soln" : None,
        },
        "OA" : {
            "slope_rho" : -0.0008958,
            "intercept_rho" : 1.02119193,
            "df_T_label" : "Acid_T",
            "initial_EV" : None,
            "initial_K" : None,
            "titration_soln" : None,
            "E0_final" : None,
            "TA_final" : None,
            "TA_processed" : None,
        }
    }

    equilibrium_constants = {
        "K_X1" : 10**-4.5,  # 10**-4.5 #midpoint pH 3 - 7,
        "K_X2" : 10**-5.25, # 10**-5.25 #midpoint pH 3 - 7.55,
        "K_X3" : 10**-5.5,   # 10**-5.5 #midpoint pH 3 - 8 (pH 8 approximate max pH)
        "carbonate" : "Lueker"
    }

    species_concentrations = {
        "BT" : 0.0004157,
        "SiT" : 0,
        "PT" : 0,
        "CTNa" : 14.999,
    }

    def set_concentrations(self,C_HCl  = 0.10060392
                               ,C_NaOH = 0.082744091
                               ,Cl_HCl = 0.10060392
                               ,I_NaOH = 0.082744091):
        self.C_HCl  = C_HCl 
        self.C_NaOH = C_NaOH
        self.Cl_HCl = Cl_HCl
        self.I_NaOH = I_NaOH
        self.titration_features["TA"]["titration_soln"] = Cl_HCl
        self.titration_features["NaOH"]["titration_soln"] = I_NaOH
        self.titration_features["OA"]["titration_soln"] = Cl_HCl


    def read_master_spreadsheet(self,master_spreadsheet_path
                               ,master_spreadsheet_filename
                               ,TA_titration_filename):

        MS_filename_full = os.path.join(master_spreadsheet_path,master_spreadsheet_filename)
        DF_MASTER = pd.read_excel(MS_filename_full)
        self.TA_filename = DF_MASTER[DF_MASTER['SAMPLE'] == TA_titration_filename].iloc[0]['SAMPLE']

        TA_IDX = DF_MASTER[DF_MASTER['SAMPLE']==self.TA_filename].index.values
        self.NaOH_filename = DF_MASTER['SAMPLE'][TA_IDX+1].item()
        self.OA_filename = DF_MASTER['SAMPLE'][TA_IDX+2].item()
        self.DF_MASTER = DF_MASTER # Might be uneccesary to keep this

        self.S_TA = DF_MASTER['SALINITY'][TA_IDX].item()
        self.V0 = DF_MASTER['g_0'][TA_IDX].item() - DF_MASTER['g_1'][TA_IDX].item()

        self.titration_features["NaOH"]["slope_rho"] = DF_MASTER['slope_NaOH'][TA_IDX].item()
        self.titration_features["NaOH"]["intercept_rho"] = DF_MASTER['intercept_NaOH'][TA_IDX].item()

        self.titration_features["TA"]["slope_rho"] = DF_MASTER['slope_HCl'][TA_IDX].item()
        self.titration_features["TA"]["intercept_rho"] = DF_MASTER['intercept_HCl'][TA_IDX].item()

        self.titration_features["OA"]["slope_rho"] = self.titration_features["OA"]["slope_rho"] 
        self.titration_features["OA"]["intercept_rho"] = self.titration_features["OA"]["intercept_rho"] 

        self.equilibrium_constants["carbonate"] = DF_MASTER['carbonate_K'][TA_IDX].item()

        self.species_concentrations['CTNa'] = DF_MASTER['CTNa'][TA_IDX].item()




    def read_excel_spreadsheets(self,TA_filename=None
                               ,NaOH_filename=None
                               ,OA_filename=None):
        # This function will read in the excel spreadsheets to memory
        # containing the organic alkalinity titration
        if TA_filename is None:
            TA_filename = self.TA_filename
        else:
            self.TA_filename = TA_filename
        if NaOH_filename is None:
            NaOH_filename = self.NaOH_filename
        else:
            self.NaOH_filename = NaOH_filename
        if OA_filename is None:
            OA_filename = self.OA_filename
        else:
            self.OA_filename = OA_filename

        TA_filename_full = os.path.join(self.dataset_path,TA_filename)
        NaOH_filename_full = os.path.join(self.dataset_path,NaOH_filename)
        OA_filename_full = os.path.join(self.dataset_path,OA_filename)

        self.df_TA = pd.read_excel(TA_filename_full)
        self.df_NaOH = pd.read_excel(NaOH_filename_full)
        self.df_OA = pd.read_excel(OA_filename_full)

    def read_dataframe(self,titration_label):
        if titration_label == "TA":
            dataframe = self.df_TA
        elif titration_label == "NaOH":
            dataframe = self.df_NaOH
        elif titration_label == "OA":
            dataframe = self.df_OA
        else:
            raise ValueError("Dataframe label not recognised")
        return dataframe

    def write_dataframe(self,dataframe,titration_label):
        if titration_label == "TA":
            self.df_TA = dataframe
        elif titration_label == "NaOH":
            self.df_NaOH = dataframe
        elif titration_label == "OA":
            self.df_OA = dataframe
        else:
            raise ValueError("Dataframe label not recognised")

    def extract_TA_data(self,g_start_idx=0,g_end_idx=9):
        # This should take the spreadsheet Dan gave me and save these data to 
        # the class instance. I've looked at it doesn't appear like the 
        # different titration classes can be all used as the same function.

        df_TA = self.df_TA
        if self.V0 is None:
            self.V0 = df_TA.iloc[g_start_idx]['g_0']-df_TA.iloc[g_start_idx]['g_1'] # Sample mass (g)
        if self.S_TA is None:
            self.S_TA = df_TA.iloc[g_start_idx]['SALINITY']  # Sample Salinity 
        self.data_start_TA = int(df_TA.iloc[g_start_idx]['data_start']-1) #row which titration starts, eg after initial acid addition and degassing
        self.titration_features["TA"]["initial_EV"] = df_TA.iloc[g_end_idx]['102 Voltage (V)'] #EV of sample before any acid addition, at index = 10

    def extract_NaOH_data(self):
        kelvin_offset = 273.15
        self.df_NaOH["T"] = (self.df_NaOH["SAMPLE Temperature (°C)"]+kelvin_offset) #create column for sample temperature (KELVIN) at each titration point
        self.df_NaOH["NaOH_T"] = self.df_NaOH["NaOH Temperature (°C)"] #create colume for temperature (Degrees Celsius) of NaOH upon addition to cell 
        self.temp_TA_known = True 

    def extract_OA_data(self,start_idx=0):

        if self.titration_features["TA"]["mass_known"] == False:
            raise AssertionError("Total Alkalinity mass must be known. Run convert_vol_to_mass on TA data first")
        if self.titration_features["NaOH"]["mass_known"] == False:
            raise AssertionError("NaOH mass must be known. Run convert_vol_to_mass on NaOH data first")

        df_TA = self.df_TA
        df_NaOH = self.df_NaOH

        self.Va = df_TA['m'][df_TA.index[-1]] #DEFINE TOTAL MASS OF ACID ADDED DURING FIRST (TA) TITRATION
        self.Vb = df_NaOH['m'][df_NaOH.index[-1]] #DEFINE TOTAL MASS OF BASE ADDED DURING NAOH TITRATION
        self.V0_OA = (self.V0+self.Va+self.Vb) # Sample mass accounting for additions of acid and base (g) 
        self.data_start_OA = int(self.df_OA.iloc[start_idx]['data_start']-1) #row which titration starts, eg after initial acid addition and degassing

    def strip_data(self,titration_label,start_idx=0,data_start=41):
        # Data start being 41 makes no sense and needs to be cleaned up into 
        # something (more?) sensible

        dataframe = self.read_dataframe(titration_label)

        if titration_label == "OA":
            data_start = self.data_start_OA

        dataframe['E(V)'] = dataframe.drop(dataframe.index[start_idx:data_start]
                                          ,axis=0)['102 Voltage (V)']
        dataframe['E(V)'] = dataframe['E(V)'].shift(-(data_start))
        dataframe['Sample_T'] = dataframe.drop(dataframe.index[start_idx:data_start]
                                              ,axis=0)['SAMPLE Temperature (°C)']
        dataframe['Sample_T'] = dataframe['Sample_T'].shift(-(data_start))
        dataframe['Acid_T'] = dataframe.drop(dataframe.index[start_idx:data_start]
                                            ,axis=0)['ACID Temperature (°C)']
        dataframe['Acid_T'] = dataframe['Acid_T'].shift(-(data_start))
        stripped_dataframe = dataframe[['E(V)', 'Sample_T', 'Acid_T', "mL"]].copy() #copy above variables to new "stripped_df"
        stripped_dataframe = stripped_dataframe.dropna() #remove NaN values from stripped_df
        
        self.write_dataframe(stripped_dataframe,titration_label)

    def vol_to_mass(self,titration_label,initial_mL=0):

        if self.temp_TA_known == False:
            raise AssertionError("NaOH temperature not known. Run extract NaOH data first.")

        slope_rho = self.titration_features[titration_label]["slope_rho"]
        intercept_rho = self.titration_features[titration_label]["intercept_rho"]

        dataframe = self.read_dataframe(titration_label)
        df_T_label = self.titration_features[titration_label]["df_T_label"]

        initial_mL = 0 if titration_label == "NaOH" else dataframe.iloc[0]["mL"] 

        dataframe["rho"] = (dataframe[df_T_label]*slope_rho+intercept_rho) # Density of titrant g.cm-3 

        initial_g = initial_mL*dataframe.iloc[0]["rho"]#convert inital volume of base/acid added to mass value (g)
        dataframe['delta_mL'] = dataframe['mL'].diff() #calculate the incremental values of volume of base/acid added
        dataframe['delta_g'] = dataframe['delta_mL']*dataframe["rho"] #convert the incremental values of volume of base/acid added to mass values
        dataframe = dataframe.fillna(0) #initial value of df['delta_g'] will be NA by default, replace with 0
        dataframe['delta_g'] = dataframe['delta_g'].fillna(0)
        dataframe['m'] = initial_g+np.cumsum(dataframe['delta_g'])#calculate cumulative total of mass of base/acid added, (initial mass of acid added + increment mass 1,  initial mass of acid added + increment mass 1 + increment mass 2 ...)

        self.titration_features[titration_label]["mass_known"] = True
        if titration_label == "TA":
            self.Va = dataframe['m'][dataframe.index[-1]]
        elif titration_label == "NaOH":
            self.Vb = dataframe['m'][dataframe.index[-1]]

        self.write_dataframe(dataframe,titration_label)

    def nernst_factor(self,titration_label):

        dataframe = self.read_dataframe(titration_label)

        dataframe["T"] = (dataframe["Sample_T"]+273.15)  # CREATE COLUMN SAMPLE TEMPERATURE (KELVIN) AT EACH TITRATION POINT
        dataframe["K"] = (self.R*dataframe['T'])/self.F # Nernst factor 
        initial_K = dataframe.iloc[9]['K'] # Initial Nernst factor, used to calculate initial pH

        self.titration_features[titration_label]["initial_K"] = initial_K

        if titration_label == "OA":
            self.titration_features[titration_label]["initial_EV"] = dataframe.iloc[0]['E(V)'] #EV of sample before any acid addition
        self.write_dataframe(dataframe,titration_label)


    def ion_strength_salinity(self,titration_label):

        dataframe = self.read_dataframe(titration_label)
        titration_soln = self.titration_features[titration_label]["titration_soln"]

        df2 = self.df_TA if titration_label == "NaOH" else self.df_NaOH
        S = self.S_TA if titration_label == "TA" else df2['S'][df2.index[-1]]

        if titration_label == "TA":
            V0 = self.V0
        elif titration_label == "NaOH":
            V0 = self.V0 + self.Va 
        elif titration_label == "OA":
            V0 = self.V0_OA

        ImO = (19.924*S/(1000-1.005*S))

        dataframe['ImO'] = (ImO*V0 + dataframe['m']*titration_soln)/(V0 + dataframe['m']) 
        dataframe['S'] = (1000*dataframe['ImO']) / (19.924 + 1.005*dataframe['ImO'])

        self.write_dataframe(dataframe,titration_label)

    def equilibrium_consts_sulfate_HF(self,titration_label):
        # Needs to be done after calculating ionic strength and salinity (same 
        # for TA and OA (similar has to be done for NaOH titration, bells &
        # whistles))
        dataframe = self.read_dataframe(titration_label)
        if titration_label == "NaOH":
            # Piss off this bridge when we get to it
            raise ValueError("We haven't implemented anything proper here yet")


        dataframe['K_S'] = np.exp(-4276.1/dataframe['T'] + 141.328 
                                  - 23.093*np.log(dataframe['T'])
                                  +(-13856/dataframe['T'] + 324.57- 47.986*np.log(dataframe['T']))*
                                  (dataframe['ImO']**(1/2)) 
                                  +(35474/dataframe['T']  - 771.54 +114.723*np.log(dataframe['T']))
                                  *dataframe['ImO'] - (2698/dataframe['T'])*(dataframe['ImO'])**(3/2)+(1776/dataframe['T'])
                                  *(dataframe['ImO'])**(2) +np.log(1-0.001005*dataframe['S'])) # pKa Bisulfate ion [HSO4-] K_S from Dickson1990


        dataframe['K_F'] = np.exp(874/dataframe['T'] - 9.68 + 0.111*dataframe['S']**0.5)  # pKa Hydrogen Fluoride ion [HF] K_F  from Perex & Fraga 1987

        dataframe['S_T'] =  (0.14/96.062)*(dataframe['S']/1.80655)# Total Sulphate concentration S_T from Morris & Riley 1966
        dataframe['F_T'] = (0.000067/18.998)*(dataframe['S']/1.80655) # Total Hydrogen Fluoride concentration F_T from Riley 1996
        dataframe["Z"] = (1+dataframe['S_T']/dataframe['K_S']) #from dickson 2007 Z = (1+S_T/K_S)

        self.write_dataframe(dataframe,titration_label)

    def gran_func(self,titration_label):

        dataframe = self.read_dataframe(titration_label)
        if titration_label == "TA":
            V0 = self.V0
        elif titration_label == "OA":
            V0 = self.V0_OA
        else:
            raise ValueError("Dataframe label not recognised")

        dataframe["F1"] = ((V0+dataframe["m"])*np.exp((dataframe["E(V)"]/(dataframe['K'])))) #Calculate Gran Function F1 at each titration point
        dataframe = dataframe[dataframe["F1"].between(10000, 1000000)] #drop all gran funciton values less than 10000 as these are TYPICALLY non linear with respect to m
        if len(dataframe.index) < 1:
            raise ValueError("No points to be fit after removing nonlinear points\nSee gran_func")
        slope, intercept, r_value, p_value, std_err =linregress(dataframe["m"], dataframe["F1"])#CALL SLOPE AND INTERCEPT OF Gran function F1
        equivalence_point = -intercept/slope #Calculate equivalence point estimate (g) from Gran function F1
        TA_est = (equivalence_point*self.C_HCl)/V0 #Estimate TA using equivalence point estimate

        dataframe["E0_est"] = (dataframe["E(V)"]-(dataframe["K"])*np.log((-V0*TA_est + dataframe["m"]*self.C_HCl)/(V0 + dataframe["m"]))) #CALCULATE EO ESTIMATE FOR EACH TITRATION POINT
        E0_init_est = dataframe["E0_est"].mean()#AVERAGE ALL EO VALUES TO OBTAIN AN INITIAL ESTIMATE OF EO
        dataframe["H"] = (np.exp((dataframe["E(V)"]-E0_init_est)/(dataframe["K"]))) #USING EO INITIAL ESTIMATE CALCULATE [H'] FOR EACH TITRATION POINT
        dataframe["GRAN_pH"] = -(np.log10(np.exp((dataframe["E(V)"]-E0_init_est)/dataframe["K"])))#CALCULATE GRAN pH

        self.titration_features[titration_label]["TA_est"]= TA_est
        self.titration_features[titration_label]["E0_init_est"]= E0_init_est
        self.write_dataframe(dataframe,titration_label)

    def nl_least_squares(self,titration_label):

        dataframe = self.read_dataframe(titration_label)
        E0_init_est = self.titration_features[titration_label]["E0_init_est"]
        TA_est = self.titration_features[titration_label]["TA_est"]
        if titration_label == "TA":
            V0 = self.V0
        elif titration_label == "OA":
            V0 = self.V0_OA
        else:
            raise ValueError("Dataframe label not recognised")

        new_dataframe = dataframe[dataframe["GRAN_pH"].between(3, 3.5)]#SELECT ONLY TITRATION POINTS WHICH pH ARE BETWEEN 3.0 - 3.5 

        # DEFINE FUNCTION WHICH RETURNS VALUE(S) TO BE MINIMISED, IN THIS CASE SSR 
        x = new_dataframe["H"]
        data = new_dataframe["m"]

        def fcn2min(params, x, data):
            f_NLSF = params['f_NLSF']
            TA_est_NLSF = params['TA_est_NLSF']
            model = ((np.sum((TA_est_NLSF + 
                              ((V0+new_dataframe["m"])/V0)*
                              ((f_NLSF*new_dataframe["H"])/new_dataframe["Z"]) -
                              (new_dataframe["m"]/V0)*self.C_HCl)**2))*10**12)

            return model - data

        # create a set of Parameters
        params = Parameters()
        params.add('f_NLSF',   value= 1)
        params.add('TA_est_NLSF', value= TA_est)
        # do fit, here with leastsq model
        minner = Minimizer(fcn2min, params, fcn_args=(x, data))
        kws  = {'options': {'maxiter':10}}
        result = minner.minimize()

        #################################
        #EXTRACT AND PROCESS TA NLSF RESULTS
        #################################
        TA_processed = result.params.get('TA_est_NLSF').value #EXTRACT INTIAL TA VALUE M/KG-1
        TA_final = TA_processed*10**6 #COVERT INTIAL TA VALUE TO µmol/KG-1
        f = 1 if titration_label == "TA" else result.params.get('f_NLSF').value  

        E0_processed = E0_init_est + dataframe["K"]*np.log(f) #CALCULATE E0 FROM NLSF F VALUE
        E0_final = E0_processed.mean() #FINAL ESTIMATE OF E0

        self.titration_features[titration_label]["E0_final"] = E0_final
        self.titration_features[titration_label]["TA_final"] = TA_final
        self.titration_features[titration_label]["TA_processed"] = TA_processed

        print(titration_label,": ", TA_final, 'µmol.kg-1')

    def pH_H_OH_H0_conc(self):
        dataframe = self.df_NaOH
        E0 = self.titration_features["TA"]["E0_final"]

        dataframe["K"] = (self.R*dataframe["T"])/self.F #Get K value at eachpoint during NaOH titration
        dataframe["pH"] = -np.log10(np.exp((dataframe["102 Voltage (V)"]-E0)/dataframe["K"]))#Using EO estimated from TA NLSF procedure calculate pH at each point during NaOH titration
        dataframe["H"] = 10**-(dataframe["pH"]) #Using pH calculate [H+] at each titration point
        dataframe['pKw'] = (-np.log10(np.exp(148.9652-13847.26/dataframe["T"] -
                           23.6521*np.log(dataframe["T"])+(-5.977+118.67/dataframe["T"] + 
                           1.0495*np.log(dataframe["T"]))*dataframe["S"]**0.5-0.01615*dataframe["S"]))) #Acid dissociation constant of Water
        dataframe['OH'] = (10**-dataframe["pKw"])/dataframe["H"] #using Acid dissociation constant of Water and [H+] calculate [OH-]
        initial_EV_NaOH = dataframe.iloc[0]['102 Voltage (V)'] #Proton concentration prior to NaOH addition, H0, defined as [H+] at end of first (TA) titration (start of NaOH titration)
        initial_K_NaOH = dataframe.iloc[0]['K']
        self.H0 = (np.exp((initial_EV_NaOH-E0)/(initial_K_NaOH)))


    def pipeline(self):
        self.set_concentrations()
        self.extract_TA_data()
        self.strip_data("TA")
        self.nernst_factor("TA")
        self.extract_NaOH_data()
        self.vol_to_mass("TA")
        self.vol_to_mass("NaOH")
        self.ion_strength_salinity("TA")
        self.equilibrium_consts_sulfate_HF("TA")
        self.gran_func("TA")
        self.nl_least_squares("TA")


        self.ion_strength_salinity("NaOH")
        self.pH_H_OH_H0_conc()

        self.extract_OA_data()
        self.strip_data("OA")
        self.vol_to_mass("OA")
        self.nernst_factor("OA")
        self.ion_strength_salinity("OA")
        self.equilibrium_consts_sulfate_HF("OA")
        self.gran_func("OA")
        self.nl_least_squares("OA")
        self.init_minimiser()
        self.dissociation_consts()


    def dissociation_consts(self,carbonate_constants=None,inc_Boron=True,inc_CTNa=False):
        dataframe = self.df_NaOH
        if carbonate_constants is None:
            carbonate_constants = self.equilibrium_constants["carbonate"]
        self.vary_CTNA = inc_CTNa

        if carbonate_constants == "Lueker":
            dataframe["pK1"] = 3633.86/dataframe["T"] - 61.2172 +9.67770*np.log(dataframe["T"]) - 0.011555*dataframe["S"] + 0.0001152*dataframe["S"]**2 
            dataframe["pK2"] = 471.78/dataframe["T"] + 25.9290 - 3.16967*np.log(dataframe["T"]) - 0.01781*dataframe["S"] + 0.0001122*dataframe["S"]**2 
            dataframe["K1"] = 10**-dataframe["pK1"] 
            dataframe["K2"] = 10**-dataframe["pK2"]
        elif carbonate_constants == "Mehrbach":
            dataframe["pK1"] = -13.7201+0.031334*dataframe["T"] + 3235.76/dataframe["T"] + (1.300*10**-5)*dataframe["S"]*dataframe["T"] - 0.1032*dataframe["S"]**0.5
            dataframe["pK2"] = 5371.9645+1.671221*dataframe["T"]+0.22913*dataframe["S"]+18.3802*np.log10(dataframe["S"])-128375.28/dataframe["T"]-2194.3055*np.log10(dataframe["T"])-(8.0944*10**-4)*dataframe["S"]*dataframe["T"]-5617.11*np.log10(dataframe["S"])/dataframe["T"] + 2.136*dataframe["S"]/dataframe["T"]
            dataframe["K1"] = 10**-dataframe["pK1"] 
            dataframe["K2"] = 10**-dataframe["pK2"]
        else:
            raise ValueError("Dissociation constants not recognised")

        dataframe['KB'] = ( (-8966.90 - 2890.53*dataframe["S"]**0.5 
                          - 77.942*dataframe["S"] + 1.728*dataframe["S"]**1.5 
                          - 0.0996*dataframe["S"]**2)/dataframe["T"]
                          + (148.0248 + 137.1942*dataframe["S"]**0.5 + 1.62142*dataframe["S"])
                          + (-24.4344-25.085*dataframe["S"]**0.5 - 0.2474*dataframe["S"])
                          *np.log(dataframe["T"])+ (0.053105*dataframe["S"]**0.5)*dataframe["T"] )

        self.species_concentrations['BT'] = inc_Boron * 0.0004157*self.S_TA/35 #TOTAL BORON [BT], (LEE2010) S VALUE IS ORIGINAL SAMPLE S
        self.species_concentrations['CTNa'] *= inc_CTNa *(self.Vb/1000)
        # Since Boron and CO2 are both false by default, I'm pretty sure that this should make sure they don't contribute 
        # unless specified.

        dataframe['KP2'] = -8814.715/dataframe['T'] + 172.0883 - 27.927*np.log(dataframe['T']) + (-160.34/dataframe['T'] + 1.3566)*dataframe['S']**0.5 + (0.37335/dataframe['T'] - 0.05778)*dataframe['S'] 
        dataframe['KSi'] = -8904.2/dataframe['T'] + 117.385 - 19.334*np.log(dataframe['T']) +(-458.79/dataframe['T'] + 3.5913)*(dataframe['ImO'])**0.5 + (188.74/dataframe['T'] - 1.5998)*(dataframe['ImO']) + (-12.1652/dataframe['T'] + 0.07871)*(dataframe['ImO'])**2 + np.log(1-0.001005*dataframe['S'])

        cleaned_dataframe = dataframe[["H", "OH", "m", "K1", "K2","pK1", "pK2",
                                       "pH" ,"KB","KP2","KSi"]].copy()

        cleaned_dataframe.dropna(inplace=True)
        cleaned_dataframe.reset_index(inplace=True)
        self.cleaned_df_NaOH = cleaned_dataframe


    def init_minimiser(self):
        self.X1 = self.titration_features["OA"]["TA_processed"]
        self.X2 = self.titration_features["OA"]["TA_processed"]
        self.X3 = self.titration_features["OA"]["TA_processed"]

        self.K_X1 = self.equilibrium_constants["K_X1"]
        self.K_X2 = self.equilibrium_constants["K_X2"]
        self.K_X3 = self.equilibrium_constants["K_X3"]

    def add_params(self,parameters,minimiser_no):
       parameters.add('CTNa',value=self.species_concentrations['CTNa'])
       if minimiser_no == 1: 
            parameters.add('H0',    value = self.H0 ) #highest [H+] value used as initial estimate
            parameters.add('C_NaOH',value = self.C_NaOH, vary = False) 
            parameters.add('X1',    value = self.X1, min = 0)
            parameters.add('K_X1',  value = self.K_X1, min = 0)
       elif minimiser_no == 2:
            parameters.add('H0',    value = self.H0, vary=False) #highest [H+] value used as initial estimate
            parameters.add('C_NaOH',value = self.C_NaOH, vary=False) 
            parameters.add('X1',    value = self.X1)
            parameters.add('K_X1',  value = self.K_X1)
            parameters.add('X2',    value = self.X2, min = 0)
            parameters.add('K_X2',  value = self.K_X2, min = 0)
       elif minimiser_no == 3:
            parameters.add('H0',    value = self.H0, vary=False) #highest [H+] value used as initial estimate
            parameters.add('C_NaOH',value = self.C_NaOH, vary=False) 
            parameters.add('X1',    value = self.X1 , vary=False)
            parameters.add('K_X1',  value = self.K_X1, vary=False)
            parameters.add('X2',    value = self.X2)
            parameters.add('K_X2',  value = self.K_X2)
            parameters.add('X3',    value = self.X3,min = 0)
            parameters.add('K_X3',  value = self.K_X3, min = 0)
       elif minimiser_no == 4:
            parameters.add('H0',    value = self.H0, vary=False) #highest [H+] value used as initial estimate
            parameters.add('C_NaOH',value = self.C_NaOH, vary=False) 
            parameters.add('X1',    value = self.X1 , vary=False)
            parameters.add('K_X1',  value = self.K_X1, vary=False)
            parameters.add('X2',    value = self.X2, vary=False)
            parameters.add('K_X2',  value = self.K_X2, vary=False)
            parameters.add('X3',    value = self.X3)
            parameters.add('K_X3',  value = self.K_X3)

    def get_params(self,result,minimiser_no):
        if minimiser_no == 1 or minimiser_no == 3:
            self.H0 = result.params.get('H0').value
        if minimiser_no == 1 or minimiser_no == 2:
            self.X1 = result.params.get('X1').value
            self.K_X1 = result.params.get('K_X1').value
            pK1 = -np.log10(self.K_X1)
        if minimiser_no == 2 or minimiser_no == 3:
            self.X2 = result.params.get('X2').value
            self.K_X2 = result.params.get('K_X2').value
            pK2 = -np.log10(self.K_X2)
        if minimiser_no == 3 or minimiser_no == 4:
            self.X3 = result.params.get('X3').value
            self.K_X3 = result.params.get('K_X3').value
            pK3 = -np.log10(self.K_X3)
        if minimiser_no == 3:
            self.C_NaOH = result.params.get('C_NaOH').value
        
        # Check the pK values are within the range of pH values for the NaOH titration
        # Throw a warning if not

        # Check that X1+X2+X3 are sensible (NB if pK[i] is dodgy X[i] is probably dodgy)


    def minimise(self,minimiser_no):

        CTNa = self.species_concentrations["CTNa"]

        if minimiser_no < 3:
            dataframe = self.cleaned_df_NaOH
        elif 2 < minimiser_no < 5:
            dataframe = self.df_NaOH
        elif minimiser_no > 4: 
            raise ValueError("minimiser_no must be in the range 1-4")

        if minimiser_no == 1:
            dataframe = dataframe[dataframe["pH"].between(0,5)]
        elif minimiser_no == 2:
            max_pH = dataframe["pH"].max()*0.8
            if max_pH < 5:
                raise ValueError("Max pH must be greater than 5")
            self.max_pH = max_pH
            dataframe = dataframe[dataframe["pH"].between(0,max_pH)]


        x = dataframe["m"]
        data = dataframe["H"]

        if minimiser_no == 1:
            def fcn2min(params, x, data):
                H0 = params['H0']
                C_NaOH = params['C_NaOH']
                X1 = params['X1']
                K_X1 = params['K_X1']

                model = ((self.V0 + self.Va+ dataframe["m"])*(dataframe["H"]-dataframe["OH"]) 
                         -((self.V0+self.Va)*H0)
                         +(dataframe["m"]*C_NaOH) 
                         - (self.V0)*(CTNa/(1+(dataframe["H"]/(dataframe["K1"]))+dataframe["K2"]/dataframe["H"]))
                         - (self.V0)*(X1/(1+dataframe["H"]/K_X1)))
                # All the inpurts going into this function are exactly the same as in the original
                # notebook. 
                return model - data
        elif minimiser_no == 2:
            def fcn2min(params, x, data):
                H0 = params['H0']
                C_NaOH = params['C_NaOH']
                X1 = params['X1']
                K_X1 = params['K_X1']
                X2 = params['X2']
                K_X2 = params['K_X2']

                model = ((self.V0 + self.Va+ dataframe["m"])*(dataframe["H"]-dataframe["OH"]) 
                         -((self.V0+self.Va)*H0)
                         +(dataframe["m"]*C_NaOH) 
                         - (self.V0)*(CTNa/(1+(dataframe["H"]/(dataframe["K1"]))+dataframe["K2"]/dataframe["H"]))
                         - (self.V0)*(X1/(1+dataframe["H"]/K_X1))
                         - (self.V0)*(X2/(1+dataframe["H"]/K_X2)))
                return model - data
        elif minimiser_no == 3:
            def fcn2min(params, x, data):
                H0 = params['H0']
                C_NaOH = params['C_NaOH']
                X1 = params['X1']
                K_X1 = params['K_X1']
                X2 = params['X2']
                K_X2 = params['K_X2']
                X3 = params['X3']
                K_X3 = params['K_X3']

                model = ((self.V0 + self.Va+ dataframe["m"])*(dataframe["H"]-dataframe["OH"]) 
                         -((self.V0+self.Va)*H0)
                         +(dataframe["m"]*C_NaOH) 
                         - (self.V0)*(CTNa/(1+(dataframe["H"]/(dataframe["K1"]))+dataframe["K2"]/dataframe["H"]))
                         - (self.V0)*(X1/(1+dataframe["H"]/K_X1))
                         - (self.V0)*(X2/(1+dataframe["H"]/K_X2))
                         - (self.V0)*(X3/(1+dataframe["H"]/K_X3)))
                return model - data
        elif minimiser_no == 4:
            def fcn2min(params, x, data):
                H0 = params['H0']
                C_NaOH = params['C_NaOH']
                X1 = params['X1']
                K_X1 = params['K_X1']
                X2 = params['X2']
                K_X2 = params['K_X2']
                X3 = params['X3']
                K_X3 = params['K_X3']
    
                model = ((self.V0 + self.Va+ dataframe["m"])*(dataframe["H"]-dataframe["OH"]) 
                         -((self.V0+self.Va)*H0)
                         +(dataframe["m"]*C_NaOH) 
                         - (self.V0)*(CTNa/(1+(dataframe["H"]/(dataframe["K1"]))+dataframe["K2"]/dataframe["H"]))
                         - (self.V0)*(X1/(1+dataframe["H"]/K_X1))
                         - (self.V0)*(X2/(1+dataframe["H"]/K_X2))
                         - (self.V0)*(X3/(1+dataframe["H"]/K_X3)))
                return model - data

        params = Parameters()
        self.add_params(params,minimiser_no)
    
        # do fit, here with leastsq model
        minner = Minimizer(fcn2min, params, fcn_args=(x, data))
        kws  = {'options': {'maxiter':100}} # Add this to minner at some point

        result = minner.minimize()
        self.get_params(result,minimiser_no)
    
    def ssr(self,minimiser_no):
        # Where do self.SiT, self.BT, self.PT come from?
        cleaned_dataframe = self.cleaned_df_NaOH
        BT = self.species_concentrations["BT"]
        PT = self.species_concentrations["PT"]
        SiT = self.species_concentrations["SiT"]
        CTNa = self.species_concentrations["CTNa"]
        dataframe = self.df_NaOH
        if minimiser_no == 1:
            cleaned_dataframe["m_calc_001"] = ( -((self.V0*(cleaned_dataframe["H"]-cleaned_dataframe["OH"]) 
                                      - self.H0*(self.V0+self.Va)
                                      - (BT*self.V0)/((cleaned_dataframe["H"]/dataframe["KB"])+1) 
                                      - (PT*self.V0)/((cleaned_dataframe["H"]/dataframe["KP2"])+1) 
                                      - (SiT*self.V0)/((cleaned_dataframe["H"]/dataframe["KSi"])+1) 
                                      - (self.X1*self.V0)/((cleaned_dataframe["H"]/self.K_X1)+1) 
                                      + self.Va*(cleaned_dataframe["H"]-cleaned_dataframe["OH"]))
                                      / ((2*CTNa)/((cleaned_dataframe["H"]**2/( dataframe["K1"]* dataframe["K2"]))+(cleaned_dataframe["H"]/dataframe["K2"])+1)
                                      + CTNa/((cleaned_dataframe["H"]/ dataframe["K1"])+(dataframe["K2"]/cleaned_dataframe["H"])+1)
                                      -cleaned_dataframe["H"] + cleaned_dataframe["OH"] + self.C_NaOH)) )
                                     
            SSR = np.sum((cleaned_dataframe['m']-cleaned_dataframe["m_calc_001"])**2)

        elif minimiser_no == 2:
            cleaned_dataframe["m_calc_002"] = ( -((self.V0*(cleaned_dataframe["H"]-cleaned_dataframe["OH"]) 
                                      - self.H0*(self.V0+self.Va)
                                      - (BT*self.V0)/((cleaned_dataframe["H"]/dataframe["KB"])+1) 
                                      - (PT*self.V0)/((cleaned_dataframe["H"]/dataframe["KP2"])+1) 
                                      - (SiT*self.V0)/((cleaned_dataframe["H"]/dataframe["KSi"])+1) 
                                      - (self.X1*self.V0)/((cleaned_dataframe["H"]/self.K_X1)+1) 
                                      - (self.X2*self.V0)/((cleaned_dataframe["H"]/self.K_X2)+1)
                                      + self.Va*(cleaned_dataframe["H"]-cleaned_dataframe["OH"]))
                                      / ((2*CTNa)/((cleaned_dataframe["H"]**2/( dataframe["K1"]* dataframe["K2"]))+(cleaned_dataframe["H"]/dataframe["K2"])+1)
                                      + CTNa/((cleaned_dataframe["H"]/ dataframe["K1"])+(dataframe["K2"]/cleaned_dataframe["H"])+1)
                                      -cleaned_dataframe["H"] + cleaned_dataframe["OH"] + self.C_NaOH)) )
                                     
            SSR = np.sum((cleaned_dataframe['m']-cleaned_dataframe["m_calc_002"])**2)

        elif minimiser_no == 3:
            cleaned_dataframe["m_calc_003"] = ( -((self.V0*(cleaned_dataframe["H"]-cleaned_dataframe["OH"]) 
                                      - self.H0*(self.V0+self.Va)
                                      - (BT*self.V0)/((cleaned_dataframe["H"]/dataframe["KB"])+1) 
                                      - (PT*self.V0)/((cleaned_dataframe["H"]/dataframe["KP2"])+1) 
                                      - (SiT*self.V0)/((cleaned_dataframe["H"]/dataframe["KSi"])+1) 
                                      - (self.X1*self.V0)/((cleaned_dataframe["H"]/self.K_X1)+1) 
                                      - (self.X2*self.V0)/((cleaned_dataframe["H"]/self.K_X2)+1)
                                      - (self.X3*self.V0)/((cleaned_dataframe["H"]/self.K_X3)+1)
                                      + self.Va*(cleaned_dataframe["H"]-cleaned_dataframe["OH"]))
                                      / ((2*CTNa)/((cleaned_dataframe["H"]**2/( dataframe["K1"]* dataframe["K2"]))+(cleaned_dataframe["H"]/dataframe["K2"])+1)
                                      + CTNa/((cleaned_dataframe["H"]/ dataframe["K1"])+(dataframe["K2"]/cleaned_dataframe["H"])+1)
                                      -cleaned_dataframe["H"] + cleaned_dataframe["OH"] + self.C_NaOH)) )
                                     
            SSR = np.sum((cleaned_dataframe['m']-cleaned_dataframe["m_calc_003"])**2)


        elif minimiser_no == 4:
            cleaned_dataframe["m_calc_004"] = ( -((self.V0*(cleaned_dataframe["H"]-cleaned_dataframe["OH"]) 
                                      - self.H0*(self.V0+self.Va)
                                      - (BT*self.V0)/((cleaned_dataframe["H"]/dataframe["KB"])+1) 
                                      - (PT*self.V0)/((cleaned_dataframe["H"]/dataframe["KP2"])+1) 
                                      - (SiT*self.V0)/((cleaned_dataframe["H"]/dataframe["KSi"])+1) 
                                      - (self.X1*self.V0)/((cleaned_dataframe["H"]/self.K_X1)+1) 
                                      - (self.X2*self.V0)/((cleaned_dataframe["H"]/self.K_X2)+1)
                                      - (self.X3*self.V0)/((cleaned_dataframe["H"]/self.K_X3)+1)
                                      + self.Va*(cleaned_dataframe["H"]-cleaned_dataframe["OH"]))
                                      / ((2*CTNa)/((cleaned_dataframe["H"]**2/( dataframe["K1"]* dataframe["K2"]))+(cleaned_dataframe["H"]/dataframe["K2"])+1)
                                      + CTNa/((cleaned_dataframe["H"]/ dataframe["K1"])+(dataframe["K2"]/cleaned_dataframe["H"])+1)
                                      -cleaned_dataframe["H"] + cleaned_dataframe["OH"] + self.C_NaOH)) )
                                     
            SSR = np.sum((cleaned_dataframe['m']-cleaned_dataframe["m_calc_004"])**2)
        return SSR 

    def repeat_minimise(self,minimiser_no,SSR_frac_change_limit=1e-4,plot_results=True):
        self.minimise(minimiser_no)
        SSR_init = self.ssr(minimiser_no)
        SSR_frac_change = 1
        num_reps = 0
        while SSR_frac_change > SSR_frac_change_limit:
            if num_reps > 100:
                raise Exception("This doesn't look like its going to work after 100 repetitions")
            self.minimise(minimiser_no)
            SSR = self.ssr(minimiser_no)
            SSR_frac_change = (((SSR  - SSR_init)/ SSR_init)**2)**0.5
            SSR_init = SSR
            num_reps += 1
        print(f"Minimisation repeated {num_reps} times in order to reach fractional change of {SSR_frac_change_limit} in SSR")
        print(f"Final SSR value = {SSR:.5f}")

        m_calc_labels = ["m_calc_001","m_calc_002","m_calc_003","m_calc_004"]

        if plot_results:
            dataframe = self.cleaned_df_NaOH if minimiser_no < 3 else self.df_NaOH

            if minimiser_no == 1:
                dataframe = dataframe[dataframe["pH"].between(0,5)]
            elif minimiser_no == 2:
                dataframe = dataframe[dataframe["pH"].between(0,self.max_pH)]

            x_meas = dataframe["m"]
            x_calc = dataframe[m_calc_labels[minimiser_no-1]]
            y_meas = dataframe["pH"]
            y_calc = dataframe["pH"]

            if minimiser_no == 1:
                print('X1 (initial):', self.X1*10**6, "| pK1(initial): ", -np.log10(self.K_X1), '| H0 :', self.H0 ) 
            elif minimiser_no == 2:
                print('X1:', self.X1*10**6, "| pK1: ", -np.log10(self.K_X1), '| Deviation % (g) NaOH :', (SSR/self.Vb)*100 ) 
            elif minimiser_no == 3:
                print('X2:', self.X2*10**6, "| pK2: ", -np.log10(self.K_X2), '| Deviation % (g) NaOH :', (SSR/self.Vb)*100 ) 
            elif minimiser_no == 4:
                print('X3:', self.X3*10**6, "| pK3: ", -np.log10(self.K_X3), '| Deviation % (g) NaOH :', (SSR/self.Vb)*100 ) 

            plt.xlabel('NaOH added (g)', fontsize=18)
            graph = plt.scatter(x_meas, y_meas, c = 'black', marker = "1")
            graph = plt.plot(x_calc, y_calc, c = 'red')
            plt.grid(False)
            ax = plt.gca()
            ax.tick_params(bottom='on', left='on', labelleft='on', labelbottom='on', length=5, labelsize = 10.5)
            plt.rc('axes',edgecolor='black')
            plt.annotate(f"SSR: {SSR:.5f}", xy=(0.0650, 0.75), xycoords='axes fraction')

            list_color  = ["black","red",]
            list_mak    = ["1",       "_"]
            list_lab    = ['Measured','Calculated',]

            ax.legend(list(zip(list_color,list_mak)), list_lab, 
                      handler_map={tuple:MarkerHandler()}) 

            plt.show()
        



    def write_to_excel(self, filename,df, sheet_name='Sheet1', startrow=None,
                           truncate_sheet=False, 
                           **to_excel_kwargs):
        """
        Append a DataFrame [df] to existing Excel file [filename]
        into [sheet_name] Sheet.
        If [filename] doesn't exist, then this function will create it.

        @param filename: File path or existing ExcelWriter
                         (Example: '/path/to/file.xlsx')
        @param df: DataFrame to save to workbook
        @param sheet_name: Name of sheet which will contain DataFrame.
                           (default: 'Sheet1')
        @param startrow: upper left cell row to dump data frame.
                         Per default (startrow=None) calculate the last row
                         in the existing DF and write to the next row...
        @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                               before writing DataFrame to Excel file
        @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                                [can be a dictionary]
        @return: None

        Usage examples:

        >>> append_df_to_excel('d:/temp/test.xlsx', df)

        >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

        >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                               index=False)

        >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2', 
                               index=False, startrow=25)

        (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
        """
        # Excel file doesn't exist - saving and exiting
        if not os.path.isfile(filename):
            df.to_excel(
                filename,
                sheet_name=sheet_name, 
                startrow=startrow if startrow is not None else 0, 
                **to_excel_kwargs)
            return

        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

        if startrow is None:
            startrow = 0

        # write out the new sheet
        df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

        # save the workbook
        writer.save()



        tit.write_to_excel("df_NaOH_output_values.xlsx",tit.df_NaOH, sheet_name='Sheet1')

    
class MarkerHandler(HandlerBase):
    def create_artists(self, legend, tup,xdescent, ydescent,
                        width, height, fontsize,trans):
        return [plt.Line2D([width/2], [height/2.],ls="",
                       marker=tup[1],color=tup[0], transform=trans)]
